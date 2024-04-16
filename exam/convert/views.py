from django.db.models import Q
from django.contrib import messages
from openpyxl import load_workbook, Workbook
from django.shortcuts import render, redirect, get_object_or_404
from .forms import UploadFileForm
from .models import ConvertConfig, Grade
from django.http import HttpResponse
from io import BytesIO
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from .forms import ConvertConfigForm, GradeForm



possible_subjects = ('语文', '数学', '数学文', '数学理', '英语', '外语', '政治', '历史', '地理', '物理', '化学',
                     '生物', '总分', '总成绩', '全科')


def sort_rule(score):
    """定义排序规则"""
    if isinstance(score, str) or score is None:
        return 0
    else:
        return float(score)


def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            # print(uploaded_file.id, uploaded_file.file.name)
            # 读取当前文件
            wb = load_workbook(file, read_only=True)
            ws = wb.active
            student_list = []
            title = []
            # 存为对象
            student_list.clear()
            for i, row in enumerate(ws.values):
                if i == 0:
                    title.extend(list(row))
                    continue
                student_list.append({'row': list(row)})
            wb.close()

            # 将数据存储到 session 中
            request.session['student_list'] = student_list
            request.session['title'] = title
            request.session['file'] = True

            # 如果使用redirect，需要将title再次存储到session中，因为redirect函数默认会将重定向的请求视为全新的请求，这意味着会创建一个新的session，而不是继续使用原始请求的session。这可能会导致在新的请求中无法访问到之前存储在session中的数据。
            request.session['title'] = request.session['title']
            request.session['student_list'] = request.session['student_list']
            request.session['file'] = True
            return redirect('index')
            # return render(request, 'convert/index.html')
    else:
        form = UploadFileForm()
    return render(request, 'convert/upload_file.html', {'form': form})


def convert_page(request):
    if request.method == 'POST':
        # 获取被选中的下拉列表的值
        selected_config_name = request.POST.get('config')  # 获取用户提交的config_name
        # 获取被选中的复选框的值
        selected_items = request.POST.getlist('selected_items')
        # 交给其他函数处理数据
        convert(request, selected_config_name, selected_items)
        return redirect(request.path)
    else:
        # 根据需要获取或使用会话中的数据
        title = request.session.get('title', [])
        # student_list = request.session.get('student_list', [])
        subjects = []
        for i, item in enumerate(title):
            if item[:2] in possible_subjects:
                subjects.append(item)

        if request.user.is_authenticated:
            # 如果用户已经登录，获取当前用户的数据和管理员的数据
            configs = ConvertConfig.objects.filter(
                Q(author=request.user) | Q(author__username='jiayezi')
            ).values_list('config_name', flat=True)
        else:
            # 如果用户没有登录，只获取jiayezi的数据
            configs = ConvertConfig.objects.filter(author__username='jiayezi').values_list('config_name', flat=True)

        context = {
            'items': subjects,
            'configs': configs,
        }
        return render(request, 'convert/convert_page.html', context)


def index(request):
    # 如果没有上传文件，就跳转到上传文件页面
    if request.session.get('file', False):
        return render(request, 'convert/index.html')
    else:
        message = '请先上传文档！'
        messages.error(request, message)  # 添加错误消息
        return redirect('upload_file')


def convert(request, config_name, selected_subject_name):
    """计算赋分成绩"""
    # 根据需要获取或使用会话中的数据
    title = request.session.get('title', [])
    student_list = request.session.get('student_list', [])

    # 查询数据库获取对应的ConvertConfig对象
    selected_config = ConvertConfig.objects.get(config_name=config_name)
    # 使用related_name获取关联的Grade对象列表
    grades = selected_config.grade_set.all()
    # 获取科目索引
    selected_subject_index = []
    for name in selected_subject_name:
        idx = title.index(name)
        selected_subject_index.append(idx)

    # 检查
    grade_info = ""
    for grade in grades:
        grade_info += f"等级: {grade.grade_name}, 高分: {grade.high_score}, 低分: {grade.low_score}, 占比: {grade.percent}\n"
    text = f'选择的科目：{", ".join(selected_subject_name)}\n'
    text += f'选择的配置：{selected_config}\n'
    text += f'配置详情：\n{grade_info}'
    print(text)

    # 加载配置文件，读取领先率、赋分区间和等级
    rate_exceed = []
    rate_dist = []
    grade_dict = {}
    rate_sum = 0
    for index, grade in enumerate(grades):
        grade_dict[index] = grade.grade_name
        rate_dist.append((grade.high_score, grade.low_score))
        rate_sum += grade.percent
        value = (100 - rate_sum) / 100.0
        rate_exceed.append(value)

    for sub_index, subject in enumerate(selected_subject_name):
        score_index = selected_subject_index[sub_index]
        student_list.sort(key=lambda x: sort_rule(x['row'][score_index]), reverse=True)

        # 获取得分大于0分的人数、获取大于0分的最小原始分
        student_data_reverse = student_list[::-1]
        student_num = len(student_data_reverse)
        min_score = 0.0
        for row_index, student in enumerate(student_data_reverse):
            if isinstance(student['row'][score_index], str) or student['row'][score_index] is None:
                continue
            if float(student['row'][score_index]) > 0.0:
                student_num -= row_index
                min_score = float(student['row'][score_index])
                break

        # 获取原始分等级区间
        rate_src = [[float(student_list[0]['row'][score_index])]]
        rate = (student_num - 1) / student_num
        previous_score = -1  # 上个分数，初始值为-1
        temp_dj = 0  # 初始等级和索引
        for row_index, student in enumerate(student_list):
            current_score_str = student['row'][score_index]
            if not isinstance(current_score_str, (int, float)):
                continue

            current_score = float(current_score_str)
            if current_score != previous_score:
                previous_score = current_score
                rate = (student_num - row_index - 1) / student_num  # 领先率
                for e_index, value in enumerate(rate_exceed):
                    # 如果这个学生的领先率大于rate_exceed里的某一个值，并且temp_dj与上次不同，就把当前分数添加到rate_src里面，然后结束内层循环。如果这个学生的领先率大于rate_exceed里的某一个值，并且temp_dj与上次相同，也要结束内层循环，所以break不能写在if内部。
                    if rate >= value:
                        if temp_dj != e_index:
                            temp_dj = e_index
                            rate_src[temp_dj - 1].append(
                                float(student_list[row_index - 1]['row'][score_index]))
                            rate_src.append([float(student['row'][score_index])])
                        break
        # rate_src[-1].append(float(student_data[-1][extra + i]))
        rate_src[-1].append(min_score)
        # print(f'{subject}原始分等级区间：{rate_src}')

        # 计算赋分成绩
        for student in student_list:
            score = student['row'][score_index]
            if not isinstance(score, (int, float)) or score == 0:
                student['row'].append('')
                student['row'].append('')
                continue
            xsdj = 0
            for index, dj_score in enumerate(rate_src):
                if dj_score[0] >= score >= dj_score[1]:
                    xsdj = index
                    break
            m = rate_src[xsdj][1]
            n = rate_src[xsdj][0]
            a = rate_dist[xsdj][1]
            b = rate_dist[xsdj][0]
            converts = (b * (score - m) + a * (n - score)) / (n - m)
            student['row'].append(grade_dict[xsdj])
            student['row'].append(round(converts))
        title.append(f'{subject}等级')
        title.append(f'{subject}赋分')

    # 将修改后的数据保存回session中
    request.session['title'] = title
    request.session['student_list'] = student_list


def rank_page(request):
    if request.method == 'POST':
        # 获取被选中的复选框的值
        selected_subjects = request.POST.getlist('selected_subjects')
        selected_groups = request.POST.getlist('selected_groups')
        # 交给其他函数处理数据
        rank(request, selected_subjects, selected_groups)
        return redirect(request.path)
    else:
        # 根据需要获取或使用会话中的数据
        title = request.session.get('title', [])
        # student_list = request.session.get('student_list', [])
        subjects = []
        groups = []
        for item in title:
            if item[:2] in possible_subjects:
                subjects.append(item)
            else:
                groups.append(item)

        context = {
            'subjects': subjects,
            'groups': groups,
        }
    return render(request, 'convert/rank_page.html', context)


def rank(request, selected_subject_name, rank_group):
    print(selected_subject_name)
    print(rank_group)


def sum_page(request):
    pass
    return render(request, 'convert/sum_page.html')


def download_file(request):
    # 获取或使用会话中的数据
    title = request.session.get('title', [])
    student_list = request.session.get('student_list', [])

    # 生成Excel文件
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.append(title)
    for student in student_list:
        ws.append(student['row'])

    # 将Excel数据写入BytesIO对象
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    # 构建HTTP响应以供下载
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="students.xlsx"'

    return response


def user_register(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')

        # 检查密码是否一致
        if password != confirm_password:
            return render(request, 'convert/register.html', {'error_message': '密码不匹配'})

        # 检查用户名是否已存在
        if User.objects.filter(username=username).exists():
            return render(request, 'convert/register.html', {'error_message': '该用户名已被注册'})

        # 创建用户
        user = User.objects.create_user(username=username, password=password)
        user.save()

        # 注册成功后
        login(request, user)
        return redirect('index')
    else:
        # 如果不是POST请求，则显示注册表单页面
        return render(request, 'convert/register.html')


def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            # 获取重定向目标，如果没有则重定向到默认页面
            redirect_to = request.GET.get('next', 'index')
            return redirect(redirect_to)
        else:
            # 登录失败的处理
            return render(request, 'convert/login.html', {'error_message': '无效的用户名或密码'})
    else:
        # 如果不是POST请求，则显示登录表单页面
        return render(request, 'convert/login.html')


def user_logout(request):
    logout(request)
    # 注销后重定向到某个页面
    return redirect('index')


@login_required
def config_list(request):
    configs = ConvertConfig.objects.filter(author=request.user)
    return render(request, 'convert/config_list.html', {'configs': configs})


@login_required
def create_config(request):
    if request.method == 'POST':
        config_form = ConvertConfigForm(request.POST)
        form_count = int(request.POST.get('form_count'))
        grade_forms = [GradeForm(request.POST, prefix=str(i)) for i in range(form_count)]
        if config_form.is_valid() and all(grade_form.is_valid() for grade_form in grade_forms):
            config = config_form.save(commit=False)
            config.author = request.user
            config.save()
            for grade_form in grade_forms:
                grade = grade_form.save(commit=False)
                grade.config_name = config
                grade.save()
            return redirect('config_list')
    else:
        config_form = ConvertConfigForm()
        grade_forms = [GradeForm(prefix=str(i)) for i in range(3)]  # 初始3个grade表单

    return render(request, 'convert/create_config.html', {'config_form': config_form, 'grade_forms': grade_forms})


@login_required
def modify_config(request, config_id):
    config = get_object_or_404(ConvertConfig, pk=config_id)
    grade_forms = [GradeForm(instance=grade) for grade in config.grade_set.all()]

    if request.method == 'POST':
        config_form = ConvertConfigForm(request.POST, instance=config)
        grade_forms = [GradeForm(request.POST, prefix=str(grade.id), instance=grade) for grade in
                       config.grade_set.all()]
        if config_form.is_valid() and all(grade_form.is_valid() for grade_form in grade_forms):
            config_form.save()
            for grade_form in grade_forms:
                grade_form.save()
            return redirect('config_list')
    else:
        config_form = ConvertConfigForm(instance=config)

    return render(request, 'convert/modify_config.html', {'config_form': config_form, 'grade_forms': grade_forms})


@login_required
def delete_config(request, config_id):
    config = get_object_or_404(ConvertConfig, pk=config_id)
    if request.method == 'POST':
        config.delete()
        return redirect('config_list')  # Replace with your success URL
    return redirect('config_list')  # Redirect back to the list view if not a POST request




# 删除数据库中的所有文件记录和文件本身
# all_files = UploadedFile.objects.all()
# for f in all_files:
#     print(f.file)
#     f.delete()
